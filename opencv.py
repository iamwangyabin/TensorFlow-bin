import os
import cv2
import numpy as np
from sklearn.model_selection import train_test_split


width_scale = 200                                       #缩放尺寸宽度
height_scale = 125                                      #缩放尺寸高度
write_path = "G:\\pytext\\reco\\datasets\\train2\\"     #要写入的图片路径
filepath = 'G:\\pytext\\reco\\datasets\\train\\'

write_path = "G:\\pytext\\reco\\datasets\\test2\\"     #要写入的图片路径
filepath = 'G:\\pytext\\reco\\datasets\\test\\'

#遍历每一张图片进行处理
def eachFile(filepath):
    i=0
    pathDir =  os.listdir(filepath)
    for allDir in pathDir:
        child = os.path.join('%s%s' % (filepath,allDir))
        write_child = os.path.join('%s%s' % (write_path,allDir))
        image = cv2.imread(child)  
        des_image=cv2.resize(image,(width_scale,height_scale),interpolation=cv2.INTER_CUBIC)
        cv2.imwrite(write_child,des_image)
        print(i)
        i+=1

eachFile(filepath)
cv2.imshow("win", des)
cv2.waitKey(0)

path = "G:\\pytext\\reco\\datasets\\train2\\"
#制作数据集
def data_label(path,count):
    data = np.empty((count,height_scale,width_scale,3),dtype = 'float32')    #建立空的四维张量类型32位浮点，最后3去掉加中1是单通道
    label = np.empty((count,),dtype = 'uint8')
    i = 0
    pathDir = os.listdir(path)
    mape={}
    fr=open('G:\\pytext\\reco\\datasets\\train.txt')
    for line in fr.readlines():
        k=line.strip().split()
        mape[k[0]]=k[1]
    for each_image in pathDir:
        all_path = os.path.join('%s%s' % (path,each_image)) #路径进行连接
        image = cv2.imread(all_path)        #后面加上0为单通道
        num = int(mape[each_image])
        array = np.asarray(image,dtype='float32')
        data[i,:,:,:] = array
        label[i] = int(num)
        i += 1
    return data,label

data,label=data_label(path,2725)
#建立交叉验证
def cross_validate(Xs, ys):
    X_train, X_test, y_train, y_test = train_test_split(
            Xs, ys, test_size=0.2, random_state=0)
    return X_train, X_test, y_train, y_test

train_data,test_data,train_label,test_label = cross_validate(data, label)


path = "G:\\pytext\\reco\\datasets\\test2\\"
def load_test(path,count):
    data = np.empty((count,height_scale,width_scale,3),dtype = 'float32')    #建立空的四维张量类型32位浮点，最后3去掉加中1是单通道
    i = 0
    mape={}
    pathDir = os.listdir(path)
    for each_image in pathDir:
        all_path = os.path.join('%s%s' % (path,each_image)) #路径进行连接
        image = cv2.imread(all_path)        #后面加上0为单通道
        array = np.asarray(image,dtype='float32')
        mape[each_image]=i
        data[i,:,:,:] = array
        i += 1
    return data,mape

test_data,mape=load_test(path,1000)



cap=cv2.VideoCapture(0)
while(True):
    ret, fram=cap.read()
    cv2.imshow('fram',fram)
    if cv2.waitKey(1) &0xFF == ord('q'):
        break


cap.release()
cv2.destroyAllWindows()


from aip import AipImageClassify

""" 你的 APPID AK SK """
APP_ID = '11217677'
API_KEY = 'OjG20o8eETyEPGaQjPKr2p1m'
SECRET_KEY = 'bgGeV5RGcHKE1Wm8vFxYfzi10uud94cT'


client = AipImageClassify(APP_ID, API_KEY, SECRET_KEY)
path = "G:\\pytext\\reco\\datasets\\test\\"

def get_file_content(filePath):
    with open(filePath, 'rb') as fp:
        return fp.read()
image = get_file_content(path)
client.logoSearch(image)


def load_test(path,mape):
    i=0
    pathDir = os.listdir(path)
    for each_image in pathDir:
        if each_image not in mape:
            all_path = os.path.join('%s%s' % (path,each_image)) 
            image = get_file_content(all_path)       
            retu=client.logoSearch(image)
            try:
                mape[each_image]=retu['result'][0]['name']
            except:
                print(each_image)
        print(i)
        i+=1


def write_test_labels(mape):
    fr=open('G:\\pytext\\reco\\testright.csv','w')
    i=0
    for root,dirs,files in os.walk('G:\\pytext\\reco\\datasets\\test'):
        for file in files:
            fr.write(file)
            fr.write(' ')
            try:
                fr.write(mape[file])
            except:
                fr.write('0')
            fr.write("\n")
            i=i+1
    fr.close()

logo={
        '卡西欧':9,
        '热风':91,
        '北面':49,
        '面包新语':85,
        '可可贝儿':11,
        '宝岛眼镜':60,
        '周大福':100,
        '背靠背':23,
        '屈臣氏':90,
        '优衣库':98,
        '佰草集':59,
        '花花公子':39,
        '吉野家':74,
        'oppo':35,
        '森马':92,
        '西贝西北菜':94,
        '汉堡王':72,
        '绝味鸭脖':75,
        '必胜客':62,
        'VERO_MODA':52,
        '谢瑞麟':97,
        '象印':96,
        '呷浦呷浦':68,
        '麦当劳':83,
        '飞亚达':67,
        '罗西尼':82,
        'New Balance':34,
        'DQ冰雪皇后':63,
        '三星':43,
        '北欧E家':61,
        '欧诗漫':87,
        '美度':32,
        '乔丹':89,
        'SELECTED':44,
        '满记甜品':84,
        '鲜芋仙':95,
        '香奈儿':10,
        'Calvin Klein':8,
        '安踏':55,
        '悦诗风吟':21,
        '欧珀莱':3,
        '梦妆':30,
        '杰克琼斯':22,
        'LACOSTE':,
        '虎牌':73,
        '稻草人箱包':65,
        '派克':37,
        '雅漾':4,
}

logok={}
for (k,v) in mape.items():
    if v in logok:
        logok[v]+=1
    else:
        logok[v]=1

dict= sorted(logok.items(), key=lambda d:d[1], reverse = True)

def write_test_labels(mape,logo):
    fr=open('G:\\pytext\\reco\\testcnn1.csv')
    new=open('G:\\pytext\\reco\\testcnn3.csv','w')
    for line in fr.readlines():
        k=line.strip().split()
        num=k[1]
        if k[0] in mape and mape[k[0]] in logo:
            num=logo[mape[k[0]]]
        new.write(k[0])
        new.write(' ')
        new.write(str(num))
        new.write("\n")
    new.close()





##以下是实验处理结果############################################################
###############################################################################
import os
import cv2
import numpy as np
import openpyxl
import matplotlib.pylab as plt
from PIL import Image


filepath='D:\\exp\\jieguo\\70.3cm R80 01\\'
write_path='D:\\exp\\final\\70.3cm R80 01\\'

def mkdir(path):  
    folder = os.path.exists(path)
    if not folder:
        os.makedirs(path)


mkdir(write_path)

D:\\exp\\2018-4-24
pathDir =  os.listdir('D:\\exp\\2018-4-26\\')
for allDir in pathDir:
    allDir=allDir.split('.cine')[0]
    child = os.path.join('%s%s' % ('D:\\exp\\jieguo\\',allDir))
    mkdir(child)


def eachFile(filepath,write_path):
    i=0
    pathDir =  os.listdir(filepath)
    for allDir in pathDir:
        child = os.path.join('%s%s' % (filepath,allDir))
        write_child = os.path.join('%s%s' % (write_path,allDir))
        image = cv2.imread(child,0)  
        cv2.imwrite(write_child,image)
        print(i)
        i+=1

def resize(filepath):
    i=0
    pathDir =  os.listdir(filepath)
    for allDir in pathDir:
        child = os.path.join('%s%s' % (filepath,allDir))
        image = Image.open(child)  
        #left up right down
        image=image.crop((270,420,550,1200))
        image.save(child)
        print(i)
        i+=1
resize('D:\\exp\\final\\70.3cm R80 01\\')

def deleteFile(write_path):
    i=0
    pathDir =  os.listdir(write_path)
    for allDir in pathDir:
        num=int(allDir.split()[1][:4])
        path = os.path.join('%s%s' % (write_path,allDir))
        if num%2==0:
            os.remove(path)
        print(i)
        i+=1


def strokeEdges(src, dst, blurKsize=5, edgeKsize=5):
    if blurKsize >= 3:
        # 7*7的模板
        blurredSrc = cv2.medianBlur(src, blurKsize)
        # BGR to GRAY
        graySrc = cv2.cvtColor(blurredSrc, cv2.COLOR_BGR2GRAY)
    else:
        graySrc = cv2.cvtColor(src, cv2.COLOR_BGR2GRAY)
    # 边缘检测
    cv2.Laplacian(graySrc, cv2.CV_8U, graySrc, ksize=edgeKsize)
    # 归一化
    normalizedInverseAlpha = (1.0 / 255) * (255 - graySrc)
    # 分割合并
    channels = cv2.split(src)
    for channel in channels:
        channel[:] = channel * normalizedInverseAlpha
    cv2.merge(channels, dst)



ix,iy=-1,-1
name=''
mape={}
def select_circle(event,x,y,flags,param):
    global circles,ix,iy,name,mape
    if event == cv2.EVENT_LBUTTONDBLCLK:
        #print(x,y)
        #write_file(x,y,name)
        for m in circles[0,:]:
            len=np.sqrt(pow((m[0]-x),2)+pow((m[1]-y),2))
            if len<3:
                cv2.circle(cimg,(m[0],m[1]),36,(0,255,255),1)
                cv2.circle(cimg,(m[0],m[1]),2,(0,0,255),3)
                mape[name]=[m[0],m[1]+36]
                print('记录圆形 '+name+' : '+str(m[0])+' '+str(m[1]+36))
    elif event == cv2.EVENT_RBUTTONDBLCLK:
        #write_file(x,y,name)
        mape[name]=[x,y]
        cv2.circle(cimg,(x,y),2,(0,0,255),3)
        print('记录右键 '+name+' : '+str(x)+' '+str(y))
    elif event == cv2.EVENT_LBUTTONDOWN:
        for m in circles[0,:]:
            if m[2]>32 and m[2]<39 and m[0]>390 and m[0]<420 :
                cv2.circle(cimg,(m[0],m[1]),36,(0,255,255),1)
                cv2.circle(cimg,(m[0],m[1]),2,(0,0,255),3)
                mape[name]=[m[0],(m[1]+36)]
        print('自动记录 '+name+' : '+str(mape[name][0])+' '+str(mape[name][1]))



def show_img(filepath):
    global circles,cimg,ix,iy,name
    pathDir =  os.listdir(filepath)
    for path in pathDir:
        img = cv2.imread(filepath+path) # queryImage  
        cimg = cv2.imread(filepath+path)
        strokeEdges(img, img, 7, 5)
        img1 = cv2.cvtColor(img,cv2.COLOR_BGR2GRAY)
        circles = cv2.HoughCircles(img1,cv2.HOUGH_GRADIENT,1,20,
                                    param1=50,param2=20,minRadius=32,maxRadius=38)
        #circles=np.around(circles)
        for i in circles[0,:]:
            if i[2]>32 and i[2]<39 and i[0]>410 and i[0]<430 :
                cv2.circle(cimg,(i[0],i[1]),36,(0,255,0),1)
                cv2.circle(cimg,(i[0],i[1]),2,(0,0,255),3)
        name=path
        cv2.imshow('image',cimg)
        cv2.setMouseCallback('image',select_circle)
        while(1):
            cv2.imshow('image',cimg)
            cv2.setMouseCallback('image',select_circle)
            k=cv2.waitKey(1)
            if k==ord('q'):
                break


show_img(filepath)
cv2.destroyAllWindows()


filepath='D:\\exp\\jieguo\\60.4cm-R360-02\\'
auto(filepath)

def auto(filepath):
    global circles,cimg,ix,iy,name,mape
    name=''
    mape={}
    pathDir =  os.listdir(filepath)
    for path in pathDir:
        img = cv2.imread(filepath+path) # queryImage  
        cimg = cv2.imread(filepath+path)
        strokeEdges(img, img, 7, 5)
        img1 = cv2.cvtColor(img,cv2.COLOR_BGR2GRAY)
        circles = cv2.HoughCircles(img1,cv2.HOUGH_GRADIENT,1,20,
                                    param1=50,param2=20,minRadius=32,maxRadius=38)
        #circles=np.around(circles)
        name=path
        highest=-1
        try:
            for i in circles[0,:]:
                if i[2]>32 and i[2]<39 and i[0]>370 and i[0]<430 :
                    cv2.circle(cimg,(i[0],i[1]),36,(0,255,0),1)
                    cv2.circle(cimg,(i[0],i[1]),2,(0,0,255),3)
                    if highest<i[1]:
                        mape[name]=[i[0],(i[1]+36)]
                        highest=i[1]
                        print(path)
        except:
            print(path)
        cv2.namedWindow('im',0)
        cv2.imshow('im',cimg)
        cv2.waitKey(1)
    cv2.destroyAllWindows()

import os
import cv2
import numpy as np
import openpyxl
import matplotlib.pylab as plt
from PIL import Image

import matplotlib.pyplot as plt
from scipy.optimize import curve_fit
import pylab


def select_point(event,x,y,flags,param):
    global circles,ix,iy,name,mape
    if event == cv2.EVENT_LBUTTONDOWN:
        cv2.circle(cimg,(x,y),2,(0,0,255),3)
        mape[name[4:8]+'up']=[x,y]
        print('记录下方点位 '+name+' : '+str(x)+' '+str(y))
    elif event == cv2.EVENT_RBUTTONDOWN:
        cv2.circle(cimg,(x,y),2,(0,0,255),3)
        mape[name[4:8]+'md']=[x,y]
        print('自由液面位置 '+name+' : '+str(x)+' '+str(y))
    elif event == cv2.EVENT_MBUTTONDOWN:  
        cv2.circle(cimg,(x,y),2,(0,0,255),3)     
        mape[name[4:8]+'down']=[x,y]
        print('颈缩位置 '+name+' : '+str(x)+' '+str(y))

def show_img(filepath):
    global cimg,ix,iy,name,mape
    mape={}
    name=''
    pathDir =  os.listdir(filepath)
    i=0
    try:
        for path in pathDir:
            if i<300 and i>10:
                i+=1
            else:
                img = cv2.imread(filepath+path) # queryImage  
                cimg = cv2.imread(filepath+path)
                name=path
                cv2.namedWindow('image',0)
                cv2.imshow('image',cimg)
                cv2.setMouseCallback('image',select_point)
                while(1):
                    cv2.namedWindow('image',0)
                    cv2.imshow('image',cimg)
                    cv2.setMouseCallback('image',select_point)
                    k=cv2.waitKey(1)
                    if k==ord('q'):
                        break
                    if k==ord('f'):
                        raise Getoutofloop()  
                i+=1
    except GeneratorExit:
        cv2.destroyAllWindows()

class Getoutofloop(Exception):
    cv2.destroyAllWindows()

filepath='D:\\exp\\jieguo\\90.6cm-R500-02\\'
show_img(filepath)
cv2.destroyAllWindows()


def write_file(path,mape):
    wb=openpyxl.Workbook()
    sheet=wb.get_sheet_by_name('Sheet')
    i=1
    for key, value in mape.items():
        sheet['A'+str(i)].value=key
        sheet['B'+str(i)].value=value[0]
        sheet['C'+str(i)].value=value[1]
        i+=1
        wb.save(path)



path='D:\\exp\\jingsuo\\90.6cm-R500-02.xlsx'
write_file(path,mape)



def write_file(path,mape):
    wb=openpyxl.Workbook()
    sheet=wb.get_sheet_by_name('Sheet')
    i=1
    for key, value in mape.items():
        sheet['A'+str(i)].value=int(key[5:8])
        sheet['B'+str(i)].value=value[0]
        sheet['C'+str(i)].value=value[1]
        i+=1
        wb.save(path)

def load_data_in_list(path):
    wb=openpyxl.load_workbook(path)
    sheet=wb.get_sheet_by_name('Sheet')
    x=[]
    y=[]
    for values in list(sheet.columns)[0]:
        x.append(values.value)
    for values in list(sheet.columns)[2]:
        y.append(values.value)
    return x,y


import matplotlib.pyplot as plt
from scipy.optimize import curve_fit
import pylab


pylab.mpl.rcParams['font.sans-serif'] = ['SimHei']

#画出某次实验结果
def draw_curve2(name,start=[]):
    path='D:\\exp\\biaoge\\'+name+'-0'
    x=[]
    y=[]
    yvals=[]
    color=['r','g','b']
    for i in range(3):
        k,j=load_data_in_list(path+str(i+1)+'auto.xlsx')
        x.append(k[start[i]:])
        y.append(j[start[i]:])
        for l in range(len(y[i])):
            y[i][l]-=j[start[i]]
            x[i][l]-=k[start[i]]
        f1 = np.polyfit(x[i], y[i], 8)  
        p1 = np.poly1d(f1)
        yvals.append(p1(x[i]))
        plt.plot(x[i], yvals[i], color[i],label='第'+str(i+1)+'次实验') 
        plt.legend(loc='upper right',fontsize=20)
    plt.show()
    return x,y,yvals

#将某次实验结果取平均值
def draw_curve(name,start=[]):
    path='D:\\exp\\biaoge\\'+name+'-0'
    x=[]
    y=[]
    yvals=[]
    cha=0.0572*72
    color=['r','g','b']
    for i in range(3):
        k,j=load_data_in_list(path+str(i+1)+'auto.xlsx')
        x.append(k[start[i]:])
        y.append(j[start[i]:])
        for l in range(len(y[i])):
            y[i][l]=(y[i][l]-j[start[i]])/72*0.0572
            x[i][l]=(x[i][l]-k[start[i]])*(1/3000)
    avx=list(set(x[1]+x[0]+x[2]))
    avx.sort()
    avy=[]
    for i in range(len(avx)):
        position=avx[i]
        c=3
        try:
            y1=y[0][x[0].index(position)]
        except:
            y1=0
            c-=1
        try:
            y2=y[1][x[1].index(position)]
        except:
            y2=0
            c-=1
        try:
            y3=y[2][x[2].index(position)]
        except:
            y3=0
            c-=1
        avy.append((y1+y2+y3)/c)
    f1 = np.polyfit(avx, avy, 8)  
    p1 = np.poly1d(f1)
    yvals=np.polyval(f1, avx) 
    #plt.plot(avx, yvals, 's', label=name ) 
    #plt.legend(loc='upper right')
    #plt.xlabel('时间 s')  
    #plt.ylabel('位移 m')  
    #plt.title('位移-时间图像')  
    #plt.show()
    return avx,f1,yvals

#只取一次实验结果
def draw_curve3(name,key=1,start=[]):
    path='D:\\exp\\biaoge\\'+name+'-0'
    x=[]
    y=[]
    yvals=[]
    cha=0.0572*72
    for i in range(3):
        k,j=load_data_in_list(path+str(i+1)+'auto.xlsx')
        x.append(k[start[i]:])
        y.append(j[start[i]:])
        for l in range(len(y[i])):
            y[i][l]=(y[i][l]-j[start[i]])/72*0.0572
            x[i][l]=(x[i][l]-k[start[i]])*(1/3000)
    avx=x[key]
    avy=[]
    for i in range(len(x[key])):
        y3=y[key][i]  
        avy.append(y3)
    f1 = np.polyfit(avx, avy, 8)  
    p1 = np.poly1d(f1)
    yvals=np.polyval(f1, avx) 
    return avx,f1,yvals


#画同一个粗糙度各个球的速度
def draw_sum(labelname):
    ax=[]
    yx=[]
    speed=[]
    name=['R500','R360','R240','R180','R80']
    color=['r','g','b','c','y']
    j,k,l=draw_curve('60.4cm-R500',start=[3,4,4])
    ax.append(j)
    yx.append(l)
    speed.append(k[:8])
    j,k,l=draw_curve3('60.4cm-R360',2,start=[3,4,2])
    ax.append(j)
    yx.append(l)
    speed.append(k[:8])
    j,k,l=draw_curve('60.4cm-R240',start=[3,5,3])
    ax.append(j)
    yx.append(l)
    speed.append(k[:8])
    j,k,l=draw_curve('60.4cm-R180',start=[3,6,2])
    ax.append(j)
    yx.append(l)
    speed.append(k[:8])
    j,k,l=draw_curve('60.4cm-R80',start=[3,3,4])
    ax.append(j)
    yx.append(l)
    speed.append(k[:8])
    for i in range(5):
        ax[i]=ax[i][:500] 
        yx[i]=yx[i][:500]
    ax[2]=ax[2][:350] 
    yx[2]=yx[2][:350]
    for i in range(5):
        for j in range(8):
            speed[i][j]=speed[i][j]*(8-j)
    yvals=[]
    for i in range(5):
        p1 = np.poly1d(speed[i])
        yvals.append(p1(ax[i]))
        plt.plot(ax[i], yvals[i], color[i],label=name[i],linewidth=2)
    plt.legend(loc='upper right',fontsize=20)
    plt.xlabel('时间/s',fontsize=20)  
    plt.ylabel('速度/(m*s-1)',fontsize=20)  
    plt.title('60cm释放高度下的速度-时间图像',fontsize=40) 
    plt.show()


#画不同粗糙度加速度
def draw_sum2():
    ax=[]
    yx=[]
    speed=[]
    name=['R500','R360','R240','R180','R80']
    color=['r','g','b','c','y']
    j,k,l=draw_curve('60.4cm-R500',start=[3,4,4])
    ax.append(j)
    yx.append(l)
    speed.append(k[:8])
    j,k,l=draw_curve3('60.4cm-R360',2,start=[3,4,2])
    ax.append(j)
    yx.append(l)
    speed.append(k[:8])
    j,k,l=draw_curve('60.4cm-R240',start=[3,5,3])
    ax.append(j)
    yx.append(l)
    speed.append(k[:8])
    j,k,l=draw_curve('60.4cm-R180',start=[3,6,2])
    ax.append(j)
    yx.append(l)
    speed.append(k[:8])
    j,k,l=draw_curve('60.4cm-R80',start=[3,3,4])
    ax.append(j)
    yx.append(l)
    speed.append(k[:8])
    for i in range(5):
        ax[i]=ax[i][:500] 
        yx[i]=yx[i][:500]
    ax[2]=ax[2][:350] 
    yx[2]=yx[2][:350]
    for i in range(5):
        for j in range(8):
            speed[i][j]=speed[i][j]*(8-j)
    for i in range(5):
        speed[i]=speed[i][:7]
        for j in range(7):
            speed[i][j]=speed[i][j]*(7-j)
    yvals=[]
    for i in range(5):
        p1 = np.poly1d(speed[i])
        yvals.append(p1(ax[i]))
        plt.plot(ax[i], yvals[i], color[i],label=name[i],linewidth=2)
    plt.legend(loc='upper right',fontsize=20)
    plt.xlabel('时间/s',fontsize=20)  
    plt.ylabel('加速度/(m*s-2)',fontsize=20)  
    plt.title('60cm释放高度下的加速度-时间图像',fontsize=40) 
    plt.show()

matplotlib.rcParams['axes.unicode_minus']=False


#不同高度下速度图
def draw_sum(labelname):
    ax=[]
    yx=[]
    speed=[]
    name=['50cm','60cm','70cm','80cm','90cm']
    color=['r','g','b','c','y']
    j,k,l=draw_curve3('49.7cm-R500',2,start=[3,4,4])
    ax.append(j)
    yx.append(l)
    speed.append(k[:8])
    j,k,l=draw_curve('60.4cm-R500',start=[3,3,5])
    ax.append(j)
    yx.append(l)
    speed.append(k[:8])
    j,k,l=draw_curve('70.3cm-R500',start=[4,5,5])
    ax.append(j)
    yx.append(l)
    speed.append(k[:8])
    j,k,l=draw_curve('80.4cm-R500',start=[3,6,4])
    ax.append(j)
    yx.append(l)
    speed.append(k[:8])
    j,k,l=draw_curve3('90.6cm-R500',2,start=[2,2,2])
    ax.append(j)
    yx.append(l)
    speed.append(k[:8])
    for i in range(5):
        ax[i]=ax[i][:500] 
        yx[i]=yx[i][:500]
    #ax[2]=ax[2][:350] 
    #yx[2]=yx[2][:350]
    for i in range(5):
        for j in range(8):
            speed[i][j]=speed[i][j]*(8-j)
    yvals=[]
    for i in range(5):
        p1 = np.poly1d(speed[i])
        yvals.append(p1(ax[i]))
        plt.plot(ax[i], yvals[i], color[i],label=name[i],linewidth=2)
    plt.legend(loc='upper right',fontsize=20)
    plt.xlabel('时间/s',fontsize=20)  
    plt.ylabel('速度/(m*s-1)',fontsize=20)  
    plt.title('不同释放高度下的速度-时间图像',fontsize=40) 
    plt.show()



def draw_sum(labelname):
    ax=[]
    yx=[]
    speed=[]
    name=['50cm','60cm','70cm','80cm','90cm']
    color=['r','g','b','c','y']
    j,k,l=draw_curve3('49.7cm-R500',2,start=[3,4,4])
    ax.append(j)
    yx.append(l)
    speed.append(k[:8])
    j,k,l=draw_curve('60.4cm-R500',start=[3,3,5])
    ax.append(j)
    yx.append(l)
    speed.append(k[:8])
    j,k,l=draw_curve('70.3cm-R500',start=[4,5,5])
    ax.append(j)
    yx.append(l)
    speed.append(k[:8])
    j,k,l=draw_curve('80.4cm-R500',start=[3,6,4])
    ax.append(j)
    yx.append(l)
    speed.append(k[:8])
    j,k,l=draw_curve3('90.6cm-R80',2,start=[5,2,4])
    ax.append(j)
    yx.append(l)
    speed.append(k[:8])
    for i in range(5):
        ax[i]=ax[i][:500] 
        yx[i]=yx[i][:500]
    for i in range(5):
        for j in range(8):
            speed[i][j]=speed[i][j]*(8-j)
    for i in range(5):
        speed[i]=speed[i][:7]
        for j in range(7):
            speed[i][j]=speed[i][j]*(7-j)
    yvals=[]
    for i in range(5):
        p1 = np.poly1d(speed[i])
        yvals.append(p1(ax[i]))
        plt.plot(ax[i], yvals[i], color[i],label=name[i],linewidth=2)
    plt.legend(loc='upper right',fontsize=20)
    plt.xlabel('时间/s',fontsize=20)  
    plt.ylabel('速度/(m*s-2)',fontsize=20)  
    plt.title('不同释放高度下的加速度-时间图像',fontsize=40) 
    plt.show()





#90.4  500
start=[[2,2,2],[1,3,2],[2,2,3],[3,3,4],[5,2,4]]
#50   80
start=[[3,3,4],[3,6,2],[3,5,3],[3,4,2,],[3,4,4]]
#60


draw_sum('labelname')

########################
##画颈缩曲线





def load_data_in_list(path):
    wb=openpyxl.load_workbook(path)
    sheet=wb.get_sheet_by_name('Sheet')
    x=[]
    y=[]
    for values in list(sheet.columns)[0]:
        x.append(int(values.value[:4]))
    for values in list(sheet.columns)[2]:
        y.append(int(values.value))
    return x,y


def draw_curve(name):
    path='D:\\exp\\jingsuo\\'+name+'.xlsx'
    x=[]
    y=[]
    yvals=[]
    x,y=load_data_in_list(path)
    for i in range(len(y)-1):
        y[i+1]=(y[i+1]-y[0])/72*0.0572
        x[i+1]=(x[i+1]-x[0])*(1/3000)
    f1 = np.polyfit(x, y, 20)  
    p1 = np.poly1d(f1)
    yvals=np.polyval(f1, x) 
    return x,f1,yvals


def draw_sum():
    pylab.mpl.rcParams['font.sans-serif'] = ['SimHei']
    ax=[]
    yx=[]
    speed=[]
    name=['R500','R360','R240','R180','R80']
    color=['r','g','b','c','y']
    j,k,l=draw_curve('49.7cm-R500-03')
    ax.append(j[1:])
    yx.append(l[1:])
    speed.append(k)
    j,k,l=draw_curve('49.7cm-R360-02')
    ax.append(j[1:])
    yx.append(l[1:])
    speed.append(k)
    j,k,l=draw_curve('49.7cm-R240-01')
    ax.append(j[1:])
    yx.append(l[1:])
    speed.append(k)
    j,k,l=draw_curve('49.7cm-R180-02')
    ax.append(j[1:])
    yx.append(l[1:])
    speed.append(k)
    j,k,l=draw_curve('49.7cm-R80-03')
    ax.append(j[1:])
    yx.append(l[1:])
    speed.append(k)
    yvals=[]
    x3=[]
    for i in range(5):
        x3.append(np.arange(ax[i][1], ax[i][-1], 0.001))
    #for i in range(5):
        #speed[i]=speed[i][:20]
        #for j in range(20):
            #speed[i][j]=speed[i][j]*(20-j)
    #for i in range(5):
        #p1 = np.poly1d(speed[i])
        #yvals.append(p1(x3))
        #plt.plot(x3, yvals[i], color[i],label=name[i],linewidth=2)
    for i in range(5):
        p1 = np.poly1d(speed[i])
        yvals.append(p1(x3[i]))
        plt.plot(x3[i], yvals[i], color[i],label=name[i],linewidth=2)
    plt.legend(loc='down right',fontsize=20)
    plt.xlabel('时间/s',fontsize=20)  
    plt.ylabel('位移/(m)',fontsize=20)  
    plt.title('不同粗糙度空泡距水面-时间图像',fontsize=40) 
    plt.show()

draw_sum()



def draw_sum():
    pylab.mpl.rcParams['font.sans-serif'] = ['SimHei']
    ax=[]
    yx=[]
    speed=[]
    name=['R500','R360','R240','R180','R80']
    color=['r','g','b','c','y']
    j,k,l=draw_curve('49.7cm-R500-03')
    ax.append(j[1:])
    yx.append(l[1:])
    speed.append(k)
    print(j[1],l[1])
    j,k,l=draw_curve('49.7cm-R360-02')
    ax.append(j[1:])
    yx.append(l[1:])
    speed.append(k)
    print(j[1],l[1])
    j,k,l=draw_curve('49.7cm-R240-01')
    ax.append(j[1:])
    yx.append(l[1:])
    speed.append(k)
    print(j[1],l[1])
    j,k,l=draw_curve('49.7cm-R180-02')
    ax.append(j[1:])
    yx.append(l[1:])
    speed.append(k)
    print(j[1],l[1])
    j,k,l=draw_curve('49.7cm-R80-03')
    ax.append(j[1:])
    yx.append(l[1:])
    speed.append(k)
    print(j[1],l[1])
    yvals=[]
    x3=[]
    for i in range(5):
        x3.append(np.arange(ax[i][1], ax[i][-1], 0.001))
    for i in range(5):
        speed[i]=speed[i][:20]
        for j in range(20):
            speed[i][j]=speed[i][j]*(20-j)
    for i in range(5):
        speed[i]=speed[i][:19]
        for j in range(19):
            speed[i][j]=speed[i][j]*(19-j)
    #for i in range(5):
        #p1 = np.poly1d(speed[i])
        #yvals.append(p1(x3))
        #plt.plot(x3, yvals[i], color[i],label=name[i],linewidth=2)
    for i in range(5):
        p1 = np.poly1d(speed[i])
        yvals.append(p1(x3[i]))
        plt.plot(x3[i], yvals[i], color[i],label=name[i],linewidth=2)
    plt.legend(loc='down right',fontsize=20)
    plt.xlabel('时间/s',fontsize=20)  
    plt.ylabel('速度/(m*s-2)',fontsize=20)  
    plt.title('不同粗糙度空泡闭合加速度-时间图像',fontsize=40) 
    plt.show()
    


draw_sum()




def draw_sum():
    pylab.mpl.rcParams['font.sans-serif'] = ['SimHei']
    ax=[]
    yx=[]
    speed=[]
    name=['50cm','60cm','70cm','80cm','90cm']
    color=['r','g','b','c','y']
    j,k,l=draw_curve('49.7cm-R500-01')
    ax.append(j[1:])
    yx.append(l[1:])
    speed.append(k)
    j,k,l=draw_curve('60.4cm-R500-02')
    ax.append(j[1:])
    yx.append(l[1:])
    speed.append(k)
    j,k,l=draw_curve('70.3cm-R500-01')
    ax.append(j[1:])
    yx.append(l[1:])
    speed.append(k)
    j,k,l=draw_curve('80.4cm-R500-02')
    ax.append(j[1:])
    yx.append(l[1:])
    speed.append(k)
    j,k,l=draw_curve('90.6cm-R500-01')
    ax.append(j[1:])
    yx.append(l[1:])
    speed.append(k)
    yvals=[]
    x3=[]
    for i in range(5):
        x3.append(np.arange(ax[i][1], 0.17, 0.001))
    for i in range(5):
        speed[i]=speed[i][:20]
        for j in range(20):
            speed[i][j]=speed[i][j]*(20-j)
    for i in range(5):
        speed[i]=speed[i][:19]
        for j in range(19):
            speed[i][j]=speed[i][j]*(19-j)
    #for i in range(5):
        #p1 = np.poly1d(speed[i])
        #yvals.append(p1(x3))
        #plt.plot(x3, yvals[i], color[i],label=name[i],linewidth=2)
    for i in range(5):
        p1 = np.poly1d(speed[i])
        yvals.append(p1(x3[i]))
        plt.plot(x3[i], yvals[i], color[i],label=name[i],linewidth=2)
    plt.legend(loc='down right',fontsize=20)
    plt.xlabel('时间/s',fontsize=20)  
    plt.ylabel('速度/(m*s-2)',fontsize=20)  
    plt.title('不同入水速度空泡闭合加速度-时间图像',fontsize=40) 
    plt.show()

draw_sum()














for i in range(5):
    plt.plot(ax[i], yx[i], color[i],label=name[i]) 


plt.xlabel('时间/s',fontsize=20)  
plt.ylabel('位移/m',fontsize=20)  
plt.title('不同粗糙度下的位移-时间图像',fontsize=20) 
plt.show()





draw_curve2('60.4cm-R360',start=[3,4,2])

x,y,yvals=draw_curve('90.6cm-R500',start=[3,3,4])

f1 = np.polyfit(ax[0], yx[0], 8)


=np.array([])